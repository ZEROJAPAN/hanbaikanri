from http.server import BaseHTTPRequestHandler
import json, io, sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def make_fill(hex6): return PatternFill(fill_type='solid', fgColor=hex6)
def make_font(bold=False, size=9, color='FF000000'): return Font(bold=bold, size=size, color=color)
def make_align(h='center', v='center', wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
thin = Side(style='thin', color='FF000000')
def make_border(): return Border(left=thin, right=thin, top=thin, bottom=thin)

FILLS = {
    'pink':   make_fill('FFFFB3C1'),
    'yellow': make_fill('FFFFFF88'),
    'lyellow':make_fill('FFFFFF00'),
    'white':  make_fill('FFFFFFFF'),
    'lblue':  make_fill('FFD0F0F8'),
    'lblue2': make_fill('FFE0E0FF'),
    'gray':   make_fill('FFC0C0C0'),
    'hdr':    make_fill('FFFFCCCC'),
    'none':   None,
}

def sc(ws, row, col, value, fill_key='white', bold=False, size=9, halign='center', border=True, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill_key and FILLS.get(fill_key): cell.fill = FILLS[fill_key]
    cell.font = make_font(bold=bold, size=size)
    cell.alignment = make_align(h=halign, v='center')
    if border: cell.border = make_border()
    if fmt: cell.number_format = fmt
    return cell

def build_nippo(store_name, date, data):
    cats = data['cats']
    daily = data['daily']
    cash_reg = data.get('cashReg') or {}
    purchase = data.get('purchase', [])
    gold_rec = data.get('goldRec', [])
    tele_rec = data.get('teleRec', [])
    stamp_rec = data.get('stampRec', [])
    visit_reasons = data.get('visitReasons', [])

    COLS = [
        {'key':'buy','chId':1,'name':'買取'},{'key':'cancel','chId':2,'name':'キャンセル'},
        {'key':'shop','chId':3,'name':'店売り'},{'key':'free','chId':4,'name':'店売り（免税）'},
        {'key':'honten','chId':5,'name':'本店'},{'key':'raku','chId':6,'name':'楽天国内'},
        {'key':'rakuEx','chId':7,'name':'楽天海外（免税）'},{'key':'yahoo','chId':8,'name':'YahooShop'},
        {'key':'yahuoku','chId':9,'name':'ヤフオク'},{'key':'wowma','chId':10,'name':'Wowma!'},
        {'key':'merukari','chId':11,'name':'メルカリ'},{'key':'ebay','chId':12,'name':'ebay（免税）'},
        {'key':'lux','chId':13,'name':'LuxUness'},{'key':'amazon','chId':14,'name':'Amazon'},
        {'key':'subscr','chId':15,'name':'サブスク'},{'key':'rakuma','chId':16,'name':'ラクマ'},
        {'key':'qoo10','chId':17,'name':'Qoo10'},{'key':'globa','chId':18,'name':'GLOBAZONE'},
        {'key':'auction','chId':19,'name':'オークション'},{'key':'web','chId':20,'name':'WEB販売'},
        {'key':'other','chId':21,'name':'その他販売'},
    ]
    TAXFREE_CH = [4, 7, 12]
    NON_SALE_CH = [1, 2]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = date

    # 列幅設定
    ws.column_dimensions['A'].width = 22
    for i in range(len(COLS)):
        ws.column_dimensions[get_column_letter(2+i*2)].width = 7
        ws.column_dimensions[get_column_letter(3+i*2)].width = 12
    ws.column_dimensions[get_column_letter(2+len(COLS)*2)].width = 14

    # データマップ
    data_map = {}
    for c in cats:
        data_map[c['id']] = {col['chId']: {'qty':0,'amt':0} for col in COLS}
    for r in daily:
        cid, chid = r.get('category_id'), r.get('channel_id')
        if cid in data_map and chid in data_map.get(cid,{}):
            data_map[cid][chid] = {'qty': r.get('quantity',0) or 0, 'amt': r.get('amount',0) or 0}

    visitors = next((r['visitors'] for r in daily if r.get('visitors',0)>0), 0)

    # 行1: 店舗名・日付
    ws.row_dimensions[1].height = 12
    ws.cell(row=1,column=1,value=store_name).font = make_font(bold=True,size=11)
    c = ws.cell(row=1,column=2+len(COLS)*2,value=date)
    c.font = make_font(bold=True,size=11)
    c.alignment = make_align(h='right')

    # 行2: チャンネル名
    ws.row_dimensions[2].height = 14.5
    sc(ws,2,1,'','hdr',bold=True,size=9)
    for i,col in enumerate(COLS):
        c1,c2 = 2+i*2, 3+i*2
        sc(ws,2,c1,col['name'],'hdr',bold=True,size=9)
        sc(ws,2,c2,'','hdr',bold=True,size=9)
        ws.merge_cells(start_row=2,start_column=c1,end_row=2,end_column=c2)

    # 行3: ヘッダー
    ws.row_dimensions[3].height = 14.5
    sc(ws,3,1,'商品名','hdr',bold=True,size=9)
    for i in range(len(COLS)):
        sc(ws,3,2+i*2,'点数','hdr',bold=True,size=9)
        sc(ws,3,3+i*2,'金額','hdr',bold=True,size=9)

    # 商品データ行
    for ci,cat in enumerate(cats):
        r = 4+ci
        ws.row_dimensions[r].height = 14.5
        sc(ws,r,1,cat['name'],'lblue',size=9,halign='left')
        for i,col in enumerate(COLS):
            d = data_map.get(cat['id'],{}).get(col['chId'],{'qty':0,'amt':0})
            sc(ws,r,2+i*2,d['qty'],'lyellow',size=9,halign='right',fmt='#,##0')
            sc(ws,r,3+i*2,d['amt'],'white',size=9,halign='right',fmt='#,##0')

    nr = 4+len(cats)

    # 計行
    ws.row_dimensions[nr].height = 14.5
    sc(ws,nr,1,'計','yellow',bold=True,size=9)
    for i,col in enumerate(COLS):
        sq = sum(data_map.get(c['id'],{}).get(col['chId'],{}).get('qty',0) for c in cats)
        sa = sum(data_map.get(c['id'],{}).get(col['chId'],{}).get('amt',0) for c in cats)
        sc(ws,nr,2+i*2,sq,'yellow',bold=True,size=9,halign='right',fmt='#,##0')
        sc(ws,nr,3+i*2,sa,'yellow',bold=True,size=9,halign='right',fmt='#,##0')
    nr+=1

    # 純売上行
    ws.row_dimensions[nr].height = 14.5
    sc(ws,nr,1,'純売上','lblue2',bold=True,size=9)
    for i,col in enumerate(COLS):
        if col['chId'] in NON_SALE_CH:
            sc(ws,nr,2+i*2,'','gray',size=9); sc(ws,nr,3+i*2,'','gray',size=9)
        else:
            amt = sum(data_map.get(c['id'],{}).get(col['chId'],{}).get('amt',0) for c in cats)
            sc(ws,nr,2+i*2,'','gray',size=9)
            sc(ws,nr,3+i*2,amt,'lblue2',bold=True,size=9,halign='right',fmt='#,##0')
    nr+=1

    # 消費税行
    ws.row_dimensions[nr].height = 14.5
    sc(ws,nr,1,'消費税','lblue2',bold=True,size=9)
    for i,col in enumerate(COLS):
        if col['chId'] in NON_SALE_CH or col['chId'] in TAXFREE_CH:
            sc(ws,nr,2+i*2,'','gray',size=9); sc(ws,nr,3+i*2,'','gray',size=9)
        else:
            amt = sum(data_map.get(c['id'],{}).get(col['chId'],{}).get('amt',0) for c in cats)
            sc(ws,nr,2+i*2,'','gray',size=9)
            sc(ws,nr,3+i*2,int(amt*10/110),'lblue2',bold=True,size=9,halign='right',fmt='#,##0')
    nr+=1

    # 現計・掛計行
    for lbl in ['現計','掛計']:
        ws.row_dimensions[nr].height = 14.5
        sc(ws,nr,1,lbl,'gray',bold=True,size=9)
        for i in range(len(COLS)):
            sc(ws,nr,2+i*2,'','gray',size=9); sc(ws,nr,3+i*2,'','gray',size=9)
        nr+=1

    # 客数行
    ws.row_dimensions[nr].height = 14.5
    sc(ws,nr,1,'客数','yellow',bold=True,size=9)
    for i,col in enumerate(COLS):
        sq = sum(data_map.get(c['id'],{}).get(col['chId'],{}).get('qty',0) for c in cats)
        sc(ws,nr,2+i*2,sq,'yellow',bold=True,size=9,halign='right',fmt='#,##0')
        sc(ws,nr,3+i*2,'','yellow',size=9)
    nr+=1

    # 来店客数
    ws.row_dimensions[nr].height = 14.5
    ws.cell(row=nr,column=7,value='来店客数').font = make_font(bold=True,size=10)
    ws.cell(row=nr,column=7).alignment = make_align(h='center')
    c = ws.cell(row=nr,column=9,value=visitors)
    c.font = make_font(bold=True,size=10); c.alignment = make_align(h='right'); c.number_format='#,##0'
    nr+=2  # 空行含む

    # 下部セクション
    sec = nr
    ws.row_dimensions[sec].height = 14.5
    ws.cell(row=sec,column=1,value='金庫・レジ').font = make_font(bold=True,size=10)
    ws.cell(row=sec,column=6,value='買取件数明細').font = make_font(bold=True,size=10)
    ws.cell(row=sec,column=13,value='地金回収').font = make_font(bold=True,size=10)
    ws.cell(row=sec,column=18,value=store_name).font = make_font(bold=True,size=9)

    hr = sec+1
    ws.row_dimensions[hr].height = 14.5
    # 金庫レジヘッダー
    sc(ws,hr,1,'','pink',bold=True,size=9); sc(ws,hr,2,'','pink',bold=True,size=9)
    sc(ws,hr,3,'レジ','pink',bold=True,size=9); sc(ws,hr,4,'','pink',bold=True,size=9)
    ws.merge_cells(start_row=hr,start_column=1,end_row=hr,end_column=2)
    ws.merge_cells(start_row=hr,start_column=3,end_row=hr,end_column=4)
    # 買取件数ヘッダー
    sc(ws,hr,6,'','pink',bold=True,size=9); sc(ws,hr,7,'','pink',bold=True,size=9)
    sc(ws,hr,8,'件数','pink',bold=True,size=9); sc(ws,hr,9,'金額','pink',bold=True,size=9)
    ws.merge_cells(start_row=hr,start_column=6,end_row=hr,end_column=7)
    # 地金回収ヘッダー
    ghdr_labels = ['','前日繰越','本日','商品化','つぶし','製品戻し','回収分','翌繰越']
    for j,lbl in enumerate(ghdr_labels):
        c1,c2 = 13+j*2, 14+j*2
        sc(ws,hr,c1,lbl,'pink',bold=True,size=9)
        sc(ws,hr,c2,'','pink',bold=True,size=9)
        if j>0: ws.merge_cells(start_row=hr,start_column=c1,end_row=hr,end_column=c2)

    cr = cash_reg
    cr_rows = [
        ('前日繰越',cr.get('carry_over',0),'pink'),
        ('本部から入金',cr.get('hq_deposit',0),'pink'),
        ('店売り（現計）',cr.get('shop_sales',0),'pink'),
        ('買取',cr.get('purchase',0),'pink'),
        ('出金',cr.get('withdrawal',0),'pink'),
        ('本日有高',cr.get('todays_amount',0),'yellow'),
        ('過不足',cr.get('shortage',0),'yellow'),
    ]

    total_v = sum(p.get('visitors',0) or 0 for p in purchase)
    total_a = sum(p.get('amount',0) or 0 for p in purchase)
    first_v = sum(p.get('first_visit',0) or 0 for p in purchase)
    repeat_v = sum(p.get('repeat_visit',0) or 0 for p in purchase)
    no_purchase = purchase[0].get('no_purchase',0) if purchase else 0
    cancel_count = purchase[0].get('cancel_count',0) if purchase else 0
    cancel_rate = (cancel_count/(total_v+cancel_count)*100) if (total_v+cancel_count)>0 else 0

    visit_rows = [
        ('初めて',first_v,0,'yellow'),
        ('２回目',repeat_v,total_a,'yellow'),
        ('小計',total_v,total_a,'yellow'),
    ]
    for vr in visit_reasons:
        p = next((x for x in purchase if x.get('visit_reason_id')==vr['id']),None)
        visit_rows.append((vr['name'], p.get('visitors',0) if p else 0, p.get('amount',0) if p else 0,'white'))
    visit_rows += [
        ('小計',total_v,total_a,'yellow'),
        ('買取不可',no_purchase,'','yellow'),
        ('キャンセル',cancel_count,'','yellow'),
        (f'キャンセル率',f'{cancel_rate:.1f}%','','yellow'),
    ]

    gold_data = [g for g in gold_rec if (g.get('today_weight',0) or 0)+(g.get('carry_over',0) or 0)>0]
    max_rows = max(len(cr_rows), len(visit_rows), len(gold_data))

    for i in range(max_rows):
        r = hr+1+i
        ws.row_dimensions[r].height = 14.5
        if i<len(cr_rows):
            lbl,val,fill = cr_rows[i]
            sc(ws,r,1,lbl,fill,bold=True,size=9,halign='left')
            sc(ws,r,2,'',fill,size=9)
            sc(ws,r,3,val,'white' if fill=='pink' else fill,bold=(fill=='yellow'),size=9,halign='right',fmt='#,##0')
            sc(ws,r,4,'','white' if fill=='pink' else fill,size=9)
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
            ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=4)
        if i<len(visit_rows):
            lbl,cnt,amt,fill = visit_rows[i]
            sc(ws,r,6,lbl,fill,bold=(fill=='yellow'),size=9,halign='left')
            sc(ws,r,7,'',fill,size=9)
            sc(ws,r,8,cnt,fill,bold=(fill=='yellow'),size=9,halign='right',fmt='#,##0')
            sc(ws,r,9,amt if amt!='' else '',fill,bold=(fill=='yellow'),size=9,halign='right',fmt='#,##0' if amt!='' else None)
            ws.merge_cells(start_row=r,start_column=6,end_row=r,end_column=7)
        if i<len(gold_data):
            g = gold_data[i]
            mtype = g.get('precious_metal_types',{})
            mname = mtype.get('name','') if isinstance(mtype,dict) else ''
            sc(ws,r,13,mname,'lblue',size=9,halign='left')
            sc(ws,r,14,'重量','lblue',size=9)
            for j,key in enumerate(['carry_over','today_weight','product_weight','crush_weight','return_weight','recovery_weight','next_carry_over']):
                c1,c2 = 15+j*2, 16+j*2
                sc(ws,r,c1,g.get(key,0) or 0,'white',size=9,halign='right',fmt='#,##0.00')
                sc(ws,r,c2,'','white',size=9)
                ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)

    # 金種表（42行目固定）
    kn_start = 42
    ws.row_dimensions[kn_start].height = 14.5
    ws.cell(row=kn_start,column=1,value='金種表').font = make_font(bold=True,size=10)
    ws.row_dimensions[kn_start+1].height = 14.5
    sc(ws,kn_start+1,1,'金種','pink',bold=True,size=9)
    sc(ws,kn_start+1,2,'枚数','pink',bold=True,size=9)
    sc(ws,kn_start+1,3,'金額','pink',bold=True,size=9)
    sc(ws,kn_start+1,4,'','pink',bold=True,size=9)
    ws.merge_cells(start_row=kn_start+1,start_column=3,end_row=kn_start+1,end_column=4)

    denoms=[('¥10,000','count_10000',10000),('¥5,000','count_5000',5000),('¥1,000','count_1000',1000),
            ('¥500','count_500',500),('¥100','count_100',100),('¥50','count_50',50),
            ('¥10','count_10',10),('¥5','count_5',5),('¥1','count_1',1)]
    for di,(name,key,val) in enumerate(denoms):
        r = kn_start+2+di
        ws.row_dimensions[r].height = 14.5
        cnt = cr.get(key,0) or 0
        sc(ws,r,1,name,'lblue',size=9,halign='right')
        sc(ws,r,2,cnt,'white',size=9,halign='right',fmt='#,##0')
        sc(ws,r,3,cnt*val,'white',size=9,halign='right',fmt='#,##0')
        sc(ws,r,4,'','white',size=9)
        ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=4)

    r = kn_start+2+len(denoms)
    ws.row_dimensions[r].height = 14.5
    sc(ws,r,1,'レジ内実現金','yellow',bold=True,size=9)
    sc(ws,r,2,'','yellow',size=9)
    sc(ws,r,3,cr.get('register_cash',0) or 0,'yellow',bold=True,size=9,halign='right',fmt='#,##0')
    sc(ws,r,4,'','yellow',size=9)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=4)
    r+=1
    ws.row_dimensions[r].height = 14.5
    sc(ws,r,1,'本部入金','pink',bold=True,size=9)
    sc(ws,r,2,cr.get('hq_payment',0) or 0,'white',size=9,halign='right',fmt='#,##0')
    sc(ws,r,3,cr.get('hq_payment',0) or 0,'white',size=9,halign='right',fmt='#,##0')
    sc(ws,r,4,'','white',size=9)
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=4)
    r+=1
    ws.row_dimensions[r].height = 14.5
    sc(ws,r,1,'レジ内繰越','yellow',bold=True,size=9)
    sc(ws,r,2,'','yellow',size=9)
    sc(ws,r,3,cr.get('register_carry',0) or 0,'yellow',bold=True,size=9,halign='right',fmt='#,##0')
    sc(ws,r,4,'','yellow',size=9)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=4)

    # テレカ・金券回収（地金の下+2行）
    if tele_rec:
        tele_col = 13
        gold_end_row = hr+1+len(gold_data)+2
        tele_r = gold_end_row
        ws.cell(row=tele_r,column=tele_col,value='テレカ・金券回収').font = make_font(bold=True,size=10)
        ws.cell(row=tele_r,column=tele_col+4,value=store_name).font = make_font(bold=True,size=9)
        tele_r+=1
        tele_hdrs=['種別','前日繰越','本日買取','回収枚数','店頭販売','翌日繰越','累計枚数']
        for j,h in enumerate(tele_hdrs):
            c1,c2 = tele_col+j*2, tele_col+j*2+1
            sc(ws,tele_r,c1,h,'pink',bold=True,size=9)
            sc(ws,tele_r,c2,'','pink',bold=True,size=9)
            ws.merge_cells(start_row=tele_r,start_column=c1,end_row=tele_r,end_column=c2)
        tele_r+=1
        for t in tele_rec:
            ttype = t.get('gift_card_types',{})
            tname = ttype.get('name','') if isinstance(ttype,dict) else ''
            sc(ws,tele_r,tele_col,tname,'lblue',size=9,halign='left')
            sc(ws,tele_r,tele_col+1,'','lblue',size=9)
            ws.merge_cells(start_row=tele_r,start_column=tele_col,end_row=tele_r,end_column=tele_col+1)
            for j,key in enumerate(['carry_over','today_purchase','recovery_count','shop_sales','next_carry_over','cumulative']):
                c1,c2 = tele_col+2+j*2, tele_col+3+j*2
                sc(ws,tele_r,c1,t.get(key,0) or 0,'white',size=9,halign='right',fmt='#,##0')
                sc(ws,tele_r,c2,'','white',size=9)
                ws.merge_cells(start_row=tele_r,start_column=c1,end_row=tele_r,end_column=c2)
            tele_r+=1
    else:
        tele_r = hr+1+len(gold_data)+2

    # 切手・印紙回収（テレカの下+2行）
    if stamp_rec:
        st_r = tele_r+1 if tele_rec else hr+1+len(gold_data)+2
        ws.cell(row=st_r,column=13,value='切手印紙回収').font = make_font(bold=True,size=10)
        ws.cell(row=st_r,column=17,value=store_name).font = make_font(bold=True,size=9)
        st_r+=1
        stamp_hdrs=['種類','回収枚数','回収金額','月末在庫枚数','月末在庫金額']
        for j,h in enumerate(stamp_hdrs):
            c1,c2 = 13+j*2, 14+j*2
            sc(ws,st_r,c1,h,'pink',bold=True,size=9)
            sc(ws,st_r,c2,'','pink',bold=True,size=9)
            ws.merge_cells(start_row=st_r,start_column=c1,end_row=st_r,end_column=c2)
        st_r+=1
        for s in stamp_rec:
            sc(ws,st_r,13,s.get('stamp_type',''),'lblue',size=9,halign='left')
            sc(ws,st_r,14,'','lblue',size=9)
            ws.merge_cells(start_row=st_r,start_column=13,end_row=st_r,end_column=14)
            for j,key in enumerate(['recovery_count','recovery_amount','stock_count','stock_amount']):
                c1,c2 = 15+j*2, 16+j*2
                sc(ws,st_r,c1,s.get(key,0) or 0,'white',size=9,halign='right',fmt='#,##0')
                sc(ws,st_r,c2,'','white',size=9)
                ws.merge_cells(start_row=st_r,start_column=c1,end_row=st_r,end_column=c2)
            st_r+=1

    # 最終行：店舗名
    ws.cell(row=ws.max_row+2,column=1,value=store_name).font = make_font(bold=True,size=10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        length = int(self.headers.get('Content-Length', 0))
        body = json.loads(self.rfile.read(length))

        store_name = body.get('storeName', '')
        date = body.get('date', '')
        data = body.get('data', {})

        xlsx_bytes = build_nippo(store_name, date, data)

        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', f'attachment; filename="nippo_{date}.xlsx"')
        self.send_header('Content-Length', str(len(xlsx_bytes)))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(xlsx_bytes)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
